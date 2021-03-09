import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


workBook = openpyxl.load_workbook ('DataSet_ExcludingDG.xlsx')
workSheet = workBook.get_sheet_by_name ('Sheet1')

newWorkBook = openpyxl.Workbook ()
newWorkSheet = newWorkBook.create_sheet ('Hourly Data')

for column in range (1, 16):
    newWorkSheet [f'{get_column_letter (column)}1'].value = workSheet [f'{get_column_letter (column)}1'].value
newWorkSheet [f'{get_column_letter (16)}1'].value = 'LastWeekLoad'
newWorkSheet [f'{get_column_letter (17)}1'].value = 'YesterdayLoad'
newWorkSheet [f'{get_column_letter (18)}1'].value = 'Load'

index_ = 0
numberOfRows = len (workSheet ['A'])
for row in range (1, numberOfRows + 1):
    oldSheetColumn = 16
    for hour in range (1, 25):
        newWorkSheet [f'A{hour + 1 + 24 * (row - 1)}'].value = index_
        for column in range (2, 16):           
            newWorkSheet [f'{get_column_letter (column)}{hour + 1 + 24 * (row - 1)}'].value = workSheet [f'{get_column_letter (column)}{row + 1}'].value
        
        newWorkSheet [f'{get_column_letter (16)}{hour + 1 + 24 * (row - 1)}'].value = workSheet [f'{get_column_letter (oldSheetColumn)}{row + 1}'].value
        newWorkSheet [f'{get_column_letter (17)}{hour + 1 + 24 * (row - 1)}'].value = workSheet [f'{get_column_letter (oldSheetColumn + 24)}{row + 1}'].value
        newWorkSheet [f'{get_column_letter (18)}{hour + 1 + 24 * (row - 1)}'].value = workSheet [f'{get_column_letter (oldSheetColumn + 48)}{row + 1}'].value
        oldSheetColumn += 1
        index_ += 1

newWorkBook.save (filename= 'DataSet_ExcludingDG_Hourly.xlsx')

input ('Please open the saved file and insert a column with header \"Hour\" in column H ... Save and close it! Then press any key to continue!')

workBook = openpyxl.load_workbook ('DataSet_ExcludingDG_Hourly.xlsx')
workSheet = workBook.get_sheet_by_name ('Hourly Data')

row = 2
while (row < len (workSheet ['B']) + 1):
    for i in range (1, 25):
        workSheet [f'H{row}'].value = i
        row += 1

workBook.save (filename= 'DataSet_ExcludingDG_Hourly.xlsx')